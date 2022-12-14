from datetime import datetime

# Como leer un archivo de excel
import openpyxl

filesheet = "./Jugadores.xlsx"

# Leer el archivo
wb = openpyxl.load_workbook(filesheet)

# Fijar la hoja
hojaPlayers = wb['Players']

#declarar matriz
tablero = []
fila=0
columna=0
vGanador=0

#rellenamos la matriz con datos
for i in range(3):
    tablero.append([" "] * 3)

def pintarTablero(tablero):
#imprimimos el tablero con los limites
    print("\n")
    print(tablero[0][0], " | ", tablero[0][1], " | ", tablero[0][2],)
    print("-------------")
    print(tablero[1][0], " | ", tablero[1][1], " | ", tablero[1][2],)
    print("-------------")
    print(tablero[2][0], " | ", tablero[2][1], " | ", tablero[2][2],)

def verificar(tablero):
    #If para validacion del jugador X
    if(tablero[0][0]=="X" and tablero[0][1]=="X"and tablero[0][2]=="X"):  
        
        vGanador=1
    elif(tablero[1][0]=="X"and tablero[1][1]=="X"and tablero[1][2]=="X"):
        
        vGanador=1
    elif(tablero[2][0]=="X"and tablero[2][1]=="X"and tablero[2][2]=="X"):
        
        vGanador=1
    elif(tablero[0][0]=="X" and tablero[1][0]=="X"and tablero[2][0]=="X"):  
        
        vGanador=1
    elif(tablero[0][1]=="X"and tablero[1][1]=="X"and tablero[2][1]=="X"):
        
        vGanador=1
    elif(tablero[0][2]=="X"and tablero[1][2]=="X"and tablero[2][2]=="X"):
        
        vGanador=1
    elif(tablero[0][0]=="X" and tablero[1][1]=="X"and tablero[2][2]=="X"):  
        
        vGanador=1
    elif(tablero[0][2]=="X"and tablero[1][1]=="X"and tablero[2][0]=="X"):
        
        vGanador=1

    #If para validacion del jugador O    
    elif(tablero[0][0]=="O" and tablero[0][1]=="O"and tablero[0][2]=="O"):  
       
        vGanador=1
    elif(tablero[1][0]=="O"and tablero[1][1]=="O"and tablero[1][2]=="O"):
        
        vGanador=1
    elif(tablero[2][0]=="O"and tablero[2][1]=="O"and tablero[2][2]=="O"):
        
        vGanador=1
    elif(tablero[0][0]=="O" and tablero[1][0]=="O"and tablero[2][0]=="O"):  
        
        vGanador=1
    elif(tablero[0][1]=="O"and tablero[1][1]=="O"and tablero[2][1]=="O"):
        
        vGanador=1
    elif(tablero[0][2]=="O"and tablero[1][2]=="O"and tablero[2][2]=="O"):
        
        vGanador=1
    elif(tablero[0][0]=="O" and tablero[1][1]=="O"and tablero[2][2]=="O"):  
        
        vGanador=1
    elif(tablero[0][2]=="O"and tablero[1][1]=="O"and tablero[2][0]=="O"):
        
        vGanador=1
    
    #If para validacion de EMPATE
    elif(tablero[0][0]=="X"and tablero[0][1]=="X"and tablero[0][2]=="O"):      
        vGanador=2
    elif(tablero[1][0]=="X"and tablero[1][1]=="O" and tablero[1][2]=="X"):        
        vGanador=2
    elif(tablero[2][0]=="O"and tablero[2][1]=="O"and tablero[2][2]=="X"):        
        vGanador=2
    elif(tablero[0][0]=="O"and tablero[0][1]=="X"and tablero[0][2]=="O"):       
        vGanador=2
    elif(tablero[1][0]=="X"and tablero[1][1]=="X" and tablero[1][2]=="O"):     
        vGanador=2
    elif(tablero[2][0]=="X"and tablero[2][1]=="O"and tablero[2][2]=="X"):        
        vGanador=2
    elif(tablero[0][0]=="O"and tablero[0][1]=="X"and tablero[0][2]=="O"):        
        vGanador=2
    elif(tablero[1][0]=="X"and tablero[1][1]=="X" and tablero[1][2]=="O"):        
        vGanador=2
    elif(tablero[2][0]=="O"and tablero[2][1]=="O"and tablero[2][2]=="X"):        
        vGanador=2
    elif(tablero[2][0]=="O"and tablero[2][1]=="O"and tablero[2][2]=="X"):        
        vGanador=2
   # elif(tablero[0][0]=="X"and tablero[0][1]=="O"and tablero[0][2]=="X" and tablero[1][0]=="X"and tablero[1][1]=="O"and tablero[1][2]=="O" and tablero[2][0]=="O"and tablero[2][1]=="X" and tablero[2][2]=="X"):
   #     vGanador=2  
   # elif(tablero[0][0]=="X"and tablero[0][1]=="O"and tablero[0][2]=="X" and tablero[1][0]=="O"and tablero[1][1]=="O"and tablero[1][2]=="X" and tablero[2][0]=="X"and tablero[2][1]=="X" and tablero[2][2]=="O"):
   #     vGanador=2    
    else: 
        vGanador=0

    return vGanador

def gatoGame():
    winDos=False
    winUno=False
    vGanador=0
    ganadorID = ""

    print("\n==========================================")
    print("\n================JUEGO GATO================")
    print("\n==========================================") 

    input("\nPresione ENTER para iniciar")

    print("\n==========SELECCION DE JUGADORES==========\n")

    #Valida cuanto jugadores existen
    for i in range(2, hojaPlayers.max_row+1):
        cantidadJugadores = i - 1

    if cantidadJugadores >= 2: 
        #Lista de jugadores disponibles
        print("Jugadores Disponibles")
        for i in range(2, hojaPlayers.max_row +1):
            print("\n")
    
            for j in range(1, hojaPlayers.max_column -4):
              celda = hojaPlayers.cell(row=i, column =j)
              print("ID:",celda.value, end = " ")

        print("\n")

        #Seleccionar jugador
        selectUno = input("Por favor, ingrese solamente el ID del jugador uno: ")
        print("")
        selectDos = input("Por favor, ingrese solamente el ID del jugador dos: ")

        #buscar en la lista
        for cell in hojaPlayers["A"]:
            if cell.value == selectUno:
                jugadorUno = hojaPlayers[f"B{cell.row}"].value #Asignamos el nombre del usuario al jugador1
                jugadorUnoID = selectUno

        for cell in hojaPlayers["A"]:
            if cell.value == selectDos:
                jugadorDos = hojaPlayers[f"B{cell.row}"].value #Asignamos el nombre del usuario al jugador2
                jugadorDosID = selectDos


        print("\n")
        print(jugadorUno," VS ", jugadorDos)
        print("\n")

        #reiniciar tablero
        tablero = []
        fila=0
        columna=0

        #rellenamos la matriz con datos
        for i in range(3):
            tablero.append([" "] * 3)

        #invocar funcion
        pintarTablero(tablero)
        repetir=True 

        while(repetir): #pasar turno hasta ganar

            if(winDos == False):
                print("\n")
                print("Turno de: ", jugadorUno)
                repeticion = 1
                while repeticion == 1:
                    fila = int(input("Coloque la fila donde quiere colcar la X:")) - 1
                    columna = int(input("Coloque la Columna donde quiere colcar la X:")) - 1
                    if tablero[fila][columna] == "X" or tablero[fila][columna] == "O":
                        print("========================================================\n")
                        print("LA POSICION YA HA SIDO SELECCIONADA ANTERIORMENTE\n")
                        print("========================================================\n")
                        repeticion = 1
                        input("PRESIONE ENTER PARA INTENTARLO DE NUEVO\n")
                    else:     
                        tablero[fila][columna] = "X"
                        pintarTablero(tablero)
                        repeticion = 0


                if (verificar(tablero) == 1):
                    ganadorID = jugadorUnoID
                    repetir = False
                    winUno = True
                    break
                #elif para identificar empate
                elif(tablero[0][0] !=' 'and tablero[0][1] !=' ' and tablero[0][2] !=' ' and tablero[1][0] !=' 'and tablero[1][1] !=' ' and tablero[1][2] !=' ' and tablero[2][0] !=' ' and tablero[2][1] !=' ' and tablero[2][2] !=' '):
                    ganadorID = 0
                    repetir = False
                    break
                elif (verificar(tablero) == 2):
                    ganadorID = 0
                    repetir = False
                    break
                    
                else:
                    print("")
            
            if (winUno == False):
                print("\n")
                print("Turno de: ", jugadorDos)
                repeticion = 1
                while repeticion == 1:
                    fila = int(input("Coloque la fila donde quiere colcar la O:")) - 1
                    columna = int(input("Coloque la Columna donde quiere colcar la O:")) - 1
                    if tablero[fila][columna] == "X" or tablero[fila][columna] == "O":
                        print("========================================================\n")
                        print("LA POSICION YA HA SIDO SELECCIONADA ANTERIORMENTE\n")
                        print("========================================================\n")
                        repeticion = 1
                        input("PRESIONE ENTER PARA INTENTARLO DE NUEVO\n")
                    else:               
                        tablero[fila][columna] = "O"
                        pintarTablero(tablero)
                        repeticion = 0
                
                    
                if (verificar(tablero) == 1):
                    ganadorID = jugadorDosID
                    repetir = False
                    winDos = True
                    break
                elif(tablero[0][0] !=' 'and tablero[0][1] !=' ' and tablero[0][2] !=' ' and tablero[1][0] !=' 'and tablero[1][1] !=' ' and tablero[1][2] !=' ' and tablero[2][0] !=' ' and tablero[2][1] !=' ' and tablero[2][2] !=' '):
                    ganadorID = 0
                    repetir = False
                    break
                #elif para identificar empate
                if (verificar(tablero) == 2):
                    ganadorID = 0
                    repetir = False
                    break
                else:
                    print("")


        #Validamos el ganador para imprimir el final correspondiente
        if(ganadorID == 0):
            print("\n============== FIN DEL JUEGO ==============\n")
            print("No hay ganador, es un EMPATE \n")     
            
        else:
            print("\n============== FIN DEL JUEGO ==============\n")
            for cell in hojaPlayers["A"]:
                if cell.value == ganadorID:
                    nombreGanador = hojaPlayers[f"B{cell.row}"].value #Asignamos el nombre del usuario al jugador1
            print("El ganador es: ", nombreGanador)
            print("\n")

        input("\nPULSE ENTER PARA CONTINUAR\n")

        #Actualizar informaci??n en excel
        i = 0
        for cell in hojaPlayers["A"]:
            i += 1
            if (cell.value == ganadorID):
                    #agregar victoria a jugador
                    ubicacionWinA = 'C'+ str(i) #colocamos la ubicaci??n de la celda de Excel
                    datoCelda = int(hojaPlayers[ubicacionWinA].value) #Convertimos el valor a INT
                    victorias = datoCelda + 1 #Sumamos valor base con el gane
                    hojaPlayers[ubicacionWinA] = str(victorias) #insertamos el nuevo valor
                        
                    #Agregar fecha actual
                    ubicacionFechaA = 'D'+ str(i) #colocamos la ubicaci??n de la celda de Excel
                    now = datetime.now()
                    fecha = now.strftime('%d/%m/%Y %H:%M') #Usamos libreria datetime para la fecha
                    hojaPlayers[ubicacionFechaA] = str(fecha) #insertamos el nuevo valor
                    wb.save(filesheet)
                    break


        print("=================== REGISTRO HISTORICO AHORCADO ===================\n")
        #Busca en la lista y se imprime los valores 
        print("ID_USUARIO         WINS         ULTIMO JUEGO")
        for select in range(2, hojaPlayers.max_row +1):
            print(hojaPlayers[f"A{select}"].value, "                ",hojaPlayers[f"C{select}"].value, "       ",hojaPlayers[f"D{select}"].value, "\n") #imprimir jugadores       
    else:
        print("No existen jugadores suficientes para jugar\n")
       
    input("PULSE ENTER PARA CONTINUAR\n")

    finalJuego()
 
def finalJuego():
    #Opciones de final del juego
        print("\n1) ??Desea volver a jugar? \n2) Volver al men?? principal \n3) Salir")
        seleccion = int(input())

        if(seleccion == 1):
            gatoGame()
        elif(seleccion == 2):
            import main
            main.menuPrincipal()
        else:
            quit()
