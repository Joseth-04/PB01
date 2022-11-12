from main import menuPrincipal, volverMenuP

# Como leer un archivo de excel
import openpyxl


filesheet = "./Jugadores.xlsx"

# Leer el archivo
wb = openpyxl.load_workbook(filesheet)

# Fijar la hoja
hojaPlayers = wb.get_sheet_by_name('Players')

def nuevoUsuario():
    #información dentro de las filas
    valido = True
    repetir = True

    #Repetir agregar
    while(repetir == True):
        #Validar que no exista el mismo ID
        while (valido == True):
            print("\n ============AGREGAR USUARIO============ \n")
            
            idUsuario = input("ID del usuario \n")

            for cell in hojaPlayers["A"]:
                if (cell.value == idUsuario):
                    valido = True
                    break
                elif(cell.value != idUsuario):
                    valido = False

            if (valido == False):
                print("ID valido")
                print ("")
                nombreUsuario = input("Nombre del usuario \n")
                datos = [(idUsuario, nombreUsuario, 0)]
                #Agregar usuario en linea
                for row in datos:
                    hojaPlayers.append(row)
                    wb.save(filesheet)
                    print("******Jugador agregado******")

            else:
                
                print("\nEl ID ingresado ya existe\nDigite uno nuevo")
            
        agregar = input("\n¿Desea agregar más usuarios? \n Digite SI o NO \n")

        if(agregar == "SI" or (agregar == "Si" or agregar == "si")):
            repetir = True
            valido = True
        else:
            repetir = False
    print()
    print ("\n 1) Menu principal \n 2) Volver")
    opcion = int(input("Ingrese la opcion a la que desea ingresar: "))
    if opcion == 1:
     menuPrincipal()
    else:
     menuJugadores()

def eliminarJugador():

    print("\n******Lista de Jugadores******")
    print("\nID", "   Nombre")
    for i in range(2, hojaPlayers.max_row +1):
        print()
        for j in range(1, hojaPlayers.max_column -3):
            celda = hojaPlayers.cell(row=i, column =j)
            print(celda.value, "  ", end = " ")
    print("")
    idUsuario = input("\nID del usuario que desea eliminar \n")
    contador = 0
    for cell in hojaPlayers["A"]:
        contador += 1
        if (cell.value == idUsuario):
            hojaPlayers.delete_rows(contador)    
            wb.save(filesheet)  
            print("******Jugador eliminado******")
            break
    print()
    print ("\n1) Menu principal",
        "\n2) Volver")
    opcion = int(input("Ingrese la opcion a la que desea ingresar: "))
    if opcion == 1:
     menuPrincipal()
    else:
     menuJugadores()
        
def mostrarJugadores():
    print("\nJugadores Disponibles")
    print("\nID", "   Nombre")
    for i in range(2, hojaPlayers.max_row +1):
        print()
        for j in range(1, hojaPlayers.max_column -3):
            celda = hojaPlayers.cell(row=i, column =j)
            print(celda.value, "  ", end = " ")
        
    print()
    print ("\n1) Menu principal",
        "\n2) Volver")
    opcion = int(input("Ingrese la opcion a la que desea ingresar: "))
    if opcion == 1:
     menuPrincipal()
    else:
     menuJugadores()

def menuJugadores():
    print ("\n1) Agregar jugador",
    "\n2) Eliminar jugador",
    "\n3) Mostrar jugadores activos",
    "\n4) Volver\n")
    opcion = int(input("Ingrese la opcion a la que desea ingresar: "))
    if opcion == 1:
        #codigo agregar jugadores    
        nuevoUsuario()
    if opcion == 2:
        #codigo agregar jugadores    
        eliminarJugador()
    if opcion == 3:
        #Código ahorcado
        mostrarJugadores()
    if opcion == 4:
        #codigo agregar jugadores    
        volverMenuP()
