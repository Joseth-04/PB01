
import getpass
from datetime import datetime

# Como leer un archivo de excel
import openpyxl


filesheet = "./Jugadores.xlsx"

# Leer el archivo
wb = openpyxl.load_workbook(filesheet)

# Fijar la hoja
hojaPlayers = wb.get_sheet_by_name('Players')

def ahorcado():

    letrasIncorrectas = ""
    #Valida cuanto jugadores existen
    for i in range(2, hojaPlayers.max_row+1):
        cantidadJugadores = i - 1

    if cantidadJugadores >= 2: 
        #Lista de jugadores disponibles
        print("Jugadores Disponibles")
        for i in range(2, hojaPlayers.max_row +1):
            print("\n")
    
            for j in range(1, hojaPlayers.max_column -4):
              celda = hojaPlayers.cell(row=i, column =j)
              print("ID:",celda.value, end = " ")

        print("\n")

        #Seleccionar jugador
        selectUno = input("Por favor, ingrese solamente el ID del jugador uno: ")
        print("")
        selectDos = input("Por favor, ingrese solamente el ID del jugador dos: ")

        #buscar en la lista
        for cell in hojaPlayers["A"]:
            if cell.value == selectUno:
                jugadorUno = hojaPlayers[f"B{cell.row}"].value #Asignamos el nombre del usuario al jugador1

        for cell in hojaPlayers["A"]:
            if cell.value == selectDos:
                jugadorDos = hojaPlayers[f"B{cell.row}"].value #Asignamos el nombre del usuario al jugador2

        print("\n", "\n")
        print(jugadorUno," VS ", jugadorDos)
        print("\n", "\n")

        input("Presione ENTER")

        print("===================================================")
        print("El jugador 1 debe ingresar la palabra oculta",
            "\nEl único caracter especial permitido es la ñ",
            "\nEl jugador 2 debe ingresar una letra",
            "\nSi la letra es correcta se marcará de lo",
            "\ncontrario se agregaráuna pieza al muñeco.",
            "\nSi logra adivinar la palabra GANA",
            "\nsino lo logra PIERDE el juego")
        print("===================================================")

        palabra = getpass.getpass("Ingrese la palabra secreta: ") #No se debe de ver
        intentos = 0
        intento = 0
        win = False
        espacios  = "_" * len(palabra)

        tableroAhorcado = ['''

        +---+
        |   |
            |
            |
            |
            |

        =========''', '''

        +----+
        |    |
        O    |
             |
             |
             |
        =========''', '''

        +----+
        |    |
        O    |
        |    |
             |
             | 
        =========''', '''

        +----+
        |    |
        O    |
       /|    |
             |
             |
        =========''', '''

        +----+
        |    |
        O    |
       /|\   |
             |
             |
        =========''', '''

        +----+
        |    |
        O    |
       /|\   |
       /     |
             |
        =========''', '''

        +----+
        |    |
        O    |
       /|\   |
       / \   |
             |
        =========''']

        while((intento < 7) or (win)):
            letraCorrecta = False
            #nuevo
            print(tableroAhorcado[intentos])
            print ("Espacios: ", espacios)

            print("\n")
            letra = input("Ingrese una letra: ")
            print("\n")

            for i in range(len(palabra)):
                if (palabra[i] in letra):
                    espacios = espacios[:i] + palabra[i] + espacios[i+1:] #Agregamos letra, borramos espacio y imprimimos guiones
                    letraCorrecta = True #Si entra una vez queda fijo y se puede validar los intentos

                #Valiamos el ganador por encontrar palabra o superar intentos
                if(espacios == palabra):
                    win = False
                    intento = 7
                    ganador = jugadorDos
                    id = selectDos
                else:
                    ganador = jugadorUno
                    id = selectUno

            if(letraCorrecta != True): #Validamos solo una vez el verdadero y sumamos intentos
                intentos += 1
                intento = intentos
                letrasIncorrectas = letrasIncorrectas + letra
                print("Letras incorrectas: ", letrasIncorrectas)
        
        #Actualizar información
        i = 0
        for cell in hojaPlayers["A"]:
            i += 1
            if (cell.value == id):
                    #agregar victoria a jugador
                    ubicacionWinA = 'E'+ str(i) #colocamos la ubicación de la celda de Excel
                    datoCelda = int(hojaPlayers[ubicacionWinA].value) #Convertimos el valor a INT
                    victorias = datoCelda + 1 #Sumamos valor base con el gane
                    hojaPlayers[ubicacionWinA] = str(victorias) #insertamos el nuevo valor
                    
                    #Agregar fecha actual
                    ubicacionFechaA = 'F'+ str(i) #colocamos la ubicación de la celda de Excel
                    now = datetime.now()
                    fecha = now.strftime('%d/%m/%Y %H:%M') #Usamos libreria datetime para la fecha
                    hojaPlayers[ubicacionFechaA] = str(fecha) #insertamos el nuevo valor
                    wb.save(filesheet)
                    break
                    
        #Validamos el ganador para imprimir el final correspondiente
        if(ganador == jugadorUno):
            print("\n============== FIN DEL JUEGO ==============\n")
            print("La palabra es: ", palabra)
            print("El ganador es: ", ganador)
            print("\n")
        else:
            print("\n============== FIN DEL JUEGO ==============\n")
            print(tableroAhorcado[intentos])
            print ("Espacios: ", espacios)
            print("El ganador es: ", ganador)
            print("\n")

        input("PULSE ENTER PARA CONTINUAR\n")
        
        #imprimir reporte
        print("\n=================== REGISTRO HISTORICO AHORCADO ===================\n")
        #buscar en la lista
        print("ID_USUARIO         WINS         ULTIMO JUEGO")
        for select in range(2, hojaPlayers.max_row +1):
            print(hojaPlayers[f"A{select}"].value, "                ",hojaPlayers[f"E{select}"].value, "       ",hojaPlayers[f"F{select}"].value, "\n") #imprimir jugadores
        
        input("PULSE ENTER PARA CONTINUAR\n")

        finalJuego()
        
    else:
        print("No existen jugadores suficientes para jugar")


def finalJuego():
    #Opciones de final del juego
        print("\n1) ¿Desea volver a jugar? \n2) Volver al menú principal \n3) Salir")
        seleccion = int(input())

        if(seleccion == 1):
            ahorcado()
        elif(seleccion == 2):
            import main
            main.menuPrincipal()
        else:
            quit()